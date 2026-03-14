"""
build_data.py — Convert To Try.xlsx to restaurants.json with geocoded coordinates.

Usage:
    python scripts/build_data.py

Reads:  To Try.xlsx (project root)
Writes: docs/data/restaurants.json
"""

import html.parser
import json
import re
import time
import urllib.request
import urllib.parse
from pathlib import Path

import openpyxl


# Manual overrides for restaurants that Nominatim can't find.
# Sourced from Google Maps.
MANUAL_COORDS: dict[str, tuple[float, float]] = {
    "White Lily Diner": (43.6686, -79.3470),        # 678 Danforth Ave
    "Bar Sugo": (43.6600, -79.4250),                 # 1162 Bloor St W (now at Geary)
    "Susie's Rise and Dine": (43.6547, -79.4108),    # 539 College St
    "Roses Social": (43.6370, -79.4210),             # Liberty Village area
    "Danny's Pizza": (43.6541, -79.4140),            # 611 College St
    "The Park LP": (43.5890, -79.6441),              # Mississauga (Port Credit)
    "Lonely Diner": (43.6555, -79.4050),             # 432 College St
    "Kabbana": (43.5550, -79.6570),                  # Mississauga (Lakeshore)
    "Sammarco": (43.6454, -79.3960),                 # Front St area
    "Enoteca Sociale": (43.6520, -79.4380),          # 1288 Dundas St W
}

PROJECT_ROOT = Path(__file__).resolve().parent.parent
# Try the copy first (original may be locked by Excel/OneDrive)
_copy = PROJECT_ROOT / "To Try_copy.xlsx"
_orig = PROJECT_ROOT / "To Try.xlsx"
XLSX_PATH = _copy if _copy.exists() else _orig
OUTPUT_PATH = PROJECT_ROOT / "docs" / "data" / "restaurants.json"

# Columns (1-indexed): B=Name, C=Type, D=Cuisine, E=Price, F=Vibe,
# G=Approximity, H=Her Rating, I=His Rating, J=Average, K=Notes, L=Status
COL_MAP = {
    "name": 2,
    "type": 3,
    "cuisine": 4,
    "price": 5,
    "vibe": 6,
    "location": 7,
    "her_rating": 8,
    "his_rating": 9,
    "notes": 11,
    "status": 12,
}


# Unsplash stock photos by cuisine — multiple per cuisine to avoid duplicates.
# Each list is cycled through so no two cards share the same image.
CUISINE_PHOTO_POOL: dict[str, list[str]] = {
    "Italian": [
        "https://images.unsplash.com/photo-1498579150354-977475b7ea0b?w=600&h=400&fit=crop",   # pasta
        "https://images.unsplash.com/photo-1595295333158-4742f28fbd85?w=600&h=400&fit=crop",   # risotto
        "https://images.unsplash.com/photo-1551183053-bf91a1d81141?w=600&h=400&fit=crop",       # bruschetta
        "https://images.unsplash.com/photo-1572441713132-c542fc4fe282?w=600&h=400&fit=crop",   # ravioli
    ],
    "Italian/French": [
        "https://images.unsplash.com/photo-1414235077428-338989a2e8c0?w=600&h=400&fit=crop",   # fine dining
        "https://images.unsplash.com/photo-1550507992-eb63ffee0847?w=600&h=400&fit=crop",       # wine & cheese
    ],
    "French": [
        "https://images.unsplash.com/photo-1555396273-367ea4eb4db5?w=600&h=400&fit=crop",       # bistro interior
        "https://images.unsplash.com/photo-1470324161839-ce2bb6fa6bc3?w=600&h=400&fit=crop",   # french cuisine
    ],
    "Steak": [
        "https://images.unsplash.com/photo-1544025162-d76694265947?w=600&h=400&fit=crop",       # steak
        "https://images.unsplash.com/photo-1558030006-450675393462?w=600&h=400&fit=crop",       # steak alt
    ],
    "Pizza": [
        "https://images.unsplash.com/photo-1565299624946-b28f40a0ae38?w=600&h=400&fit=crop",
    ],
    "Pancakes": [
        "https://images.unsplash.com/photo-1567620905732-2d1ec7ab7445?w=600&h=400&fit=crop",
    ],
    "International": [
        "https://images.unsplash.com/photo-1504674900247-0877df9cc836?w=600&h=400&fit=crop",   # spread
        "https://images.unsplash.com/photo-1476224203421-9ac39bcb3327?w=600&h=400&fit=crop",   # plates
    ],
    "international": [
        "https://images.unsplash.com/photo-1466978913421-dad2ebd01d17?w=600&h=400&fit=crop",   # table setting
        "https://images.unsplash.com/photo-1529543544282-ea96407407c3?w=600&h=400&fit=crop",   # shared plates
    ],
    "Drinks": [
        "https://images.unsplash.com/photo-1514362545857-3bc16c4c7d1b?w=600&h=400&fit=crop",
    ],
    "Mediterainian": [
        "https://images.unsplash.com/photo-1544124499-58912cbddaad?w=600&h=400&fit=crop",
    ],
    "Italian Sandwich": [
        "https://images.unsplash.com/photo-1509722747041-616f39b57569?w=600&h=400&fit=crop",
    ],
}
FALLBACK_PHOTOS = [
    "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=600&h=400&fit=crop",
    "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=600&h=400&fit=crop",
    "https://images.unsplash.com/photo-1537047902294-62a40c20a6ae?w=600&h=400&fit=crop",
]

# Track which photos have been used globally to avoid duplicates
_used_photos: set[str] = set()


def pick_stock_photo(cuisine: str) -> str:
    """Pick a stock photo for the cuisine, ensuring no duplicates across all cards."""
    pool = CUISINE_PHOTO_POOL.get(cuisine, FALLBACK_PHOTOS)
    # Pick the first unused photo from the pool
    for photo in pool:
        if photo not in _used_photos:
            _used_photos.add(photo)
            return photo
    # All pool photos used — try fallbacks
    for photo in FALLBACK_PHOTOS:
        if photo not in _used_photos:
            _used_photos.add(photo)
            return photo
    # Last resort: reuse first from pool
    return pool[0]


class OGImageParser(html.parser.HTMLParser):
    """Extract og:image from HTML <meta> tags."""

    def __init__(self):
        super().__init__()
        self.og_image = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag != "meta":
            return
        attr_dict = {k.lower(): v for k, v in attrs if v}
        prop = attr_dict.get("property", "").lower()
        if prop == "og:image" and "content" in attr_dict:
            self.og_image = attr_dict["content"]


def fetch_og_image(url: str) -> str | None:
    """Try to fetch the og:image from a URL's HTML."""
    if not url:
        return None
    # Skip PDFs and non-HTML
    if url.endswith(".pdf"):
        return None
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (compatible; TorontoEats/1.0)",
        })
        with urllib.request.urlopen(req, timeout=8) as resp:
            content_type = resp.headers.get("Content-Type", "")
            if "html" not in content_type.lower():
                return None
            body = resp.read(50_000).decode("utf-8", errors="ignore")
            parser = OGImageParser()
            parser.feed(body)
            if parser.og_image and parser.og_image.startswith("http"):
                return parser.og_image
    except Exception:
        pass
    return None


def get_image(name: str, url: str, cuisine: str) -> tuple[str, str]:
    """Get an image for a restaurant. Returns (url, type) where type is 'og', 'instagram', or 'stock'."""
    print(f"    image: ", end="")
    og = fetch_og_image(url)
    if og:
        if "instagram" in (url or ""):
            print("og:image from instagram (thumbnail)")
            return og, "instagram"
        print("og:image from site")
        return og, "og"
    fallback = pick_stock_photo(cuisine)
    print(f"cuisine fallback ({cuisine})")
    return fallback, "stock"


def geocode_nominatim(query: str) -> tuple[float | None, float | None]:
    """Geocode using OpenStreetMap Nominatim (free, no key)."""
    url = (
        "https://nominatim.openstreetmap.org/search?"
        + urllib.parse.urlencode({"q": query, "format": "json", "limit": 1})
    )
    req = urllib.request.Request(url, headers={"User-Agent": "TorontoRestaurants/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
            if data:
                return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception as exc:
        print(f"    nominatim error: {exc}")
    return None, None


def geocode(name: str, location: str) -> tuple[float | None, float | None]:
    """Try multiple geocoding strategies to find the restaurant."""
    strategies = [
        f"{name}, {location}, Toronto, Ontario, Canada",
        f"{name}, Toronto, Ontario, Canada",
        f"{name} restaurant, Toronto, Canada",
    ]
    for query in strategies:
        lat, lng = geocode_nominatim(query)
        if lat is not None:
            return lat, lng
        time.sleep(1.1)
    return None, None


def read_restaurants() -> list[dict]:
    """Read restaurant rows from the xlsx file."""
    # Load twice: data_only for values, regular for hyperlinks
    wb_data = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    wb_links = openpyxl.load_workbook(XLSX_PATH)
    ws = wb_data["Restaurants"]
    ws_links = wb_links["Restaurants"]

    restaurants = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, COL_MAP["name"]).value
        if not name or not str(name).strip():
            continue

        her = ws.cell(row, COL_MAP["her_rating"]).value
        his = ws.cell(row, COL_MAP["his_rating"]).value
        her_val = float(her) if her is not None else None
        his_val = float(his) if his is not None else None

        if her_val is not None and his_val is not None:
            avg = round((her_val + his_val) / 2, 1)
        elif her_val is not None:
            avg = her_val
        elif his_val is not None:
            avg = his_val
        else:
            avg = None

        status_raw = ws.cell(row, COL_MAP["status"]).value
        status = status_raw.strip() if status_raw else "Want to try"

        # Extract hyperlink from name cell
        link_cell = ws_links.cell(row, COL_MAP["name"])
        url = link_cell.hyperlink.target if link_cell.hyperlink else ""
        # Clean tracking params from URLs
        if url and "?" in url:
            base = url.split("?")[0]
            # Keep query params only if they look intentional (not fbclid/utm)
            query = url.split("?", 1)[1]
            if any(p in query for p in ["fbclid", "utm_", "PAZXh0"]):
                url = base

        restaurants.append({
            "name": str(name).strip(),
            "url": url,
            "type": str(ws.cell(row, COL_MAP["type"]).value or "").strip(),
            "cuisine": str(ws.cell(row, COL_MAP["cuisine"]).value or "").strip(),
            "price": str(ws.cell(row, COL_MAP["price"]).value or "").strip(),
            "vibe": str(ws.cell(row, COL_MAP["vibe"]).value or "").strip(),
            "location": str(ws.cell(row, COL_MAP["location"]).value or "").strip(),
            "her_rating": her_val,
            "his_rating": his_val,
            "average_rating": avg,
            "notes": str(ws.cell(row, COL_MAP["notes"]).value or "").strip(),
            "status": status,
            "lat": None,
            "lng": None,
            "image": "",
            "image_type": "",
        })

    return restaurants


def main() -> None:
    print("Reading spreadsheet...")
    restaurants = read_restaurants()
    print(f"Found {len(restaurants)} restaurants.\n")

    print("Geocoding locations...")
    for rest in restaurants:
        print(f"  {rest['name']} ({rest['location']})...", end=" ")

        # Check manual overrides first
        if rest["name"] in MANUAL_COORDS:
            lat, lng = MANUAL_COORDS[rest["name"]]
            print(f"({lat}, {lng}) [manual]")
        else:
            lat, lng = geocode(rest["name"], rest["location"])
            status = f"({lat}, {lng})" if lat else "FAILED"
            print(status)

        rest["lat"] = lat
        rest["lng"] = lng

    print("\nFetching images...")
    for rest in restaurants:
        print(f"  {rest['name']}...", end=" ")
        img_url, img_type = get_image(rest["name"], rest["url"], rest["cuisine"])
        rest["image"] = img_url
        rest["image_type"] = img_type
        time.sleep(0.3)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(restaurants, f, indent=2, ensure_ascii=False)

    succeeded = sum(1 for r in restaurants if r["lat"] is not None)
    print(f"\nDone! {succeeded}/{len(restaurants)} geocoded.")
    print(f"Output: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
